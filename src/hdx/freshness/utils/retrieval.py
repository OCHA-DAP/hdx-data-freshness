"""Utility to download and hash resources. Uses asyncio. Note that the purpose of
asyncio is to help with IO-bound rather than CPU-bound code (for which multiprocessing
is more suitable as it leverages multiple CPUs). Asyncio allows you to structure your
code so that when one piece of linear single-threaded code (coroutine) is waiting for
something to happen another can take over and use the CPU. While conceptually similar to
threading, the difference is that with asyncio, it is the task of the developer rather
than the OS to decide when to switch to the next task.
"""
import asyncio
import hashlib
import logging
from io import BytesIO
from timeit import default_timer as timer
from typing import Dict, List, Optional, Tuple, Union

import aiohttp
import tqdm
import uvloop
from dateutil import parser
from openpyxl import load_workbook

from . import retry
from .ratelimiter import RateLimiter

logger = logging.getLogger(__name__)


class Retrieval:
    """Retrieval class for downloading and hashing resources.

    Args:
        user_agent (str): User agent string to use when downloading
        url_ignore (Optional[str]): Parts of url to ignore for special xlsx handling
    """

    toolargeerror = "File too large to hash!"
    notmatcherror = "does not match HDX format"
    clienterror_regex = ".Client(.*)Error "
    ignore_mimetypes = ["application/octet-stream", "application/binary"]
    mimetypes = {
        "json": ["application/json"],
        "geojson": ["application/json", "application/geo+json"],
        "shp": ["application/zip", "application/x-zip-compressed"],
        "csv": ["text/csv", "application/zip", "application/x-zip-compressed"],
        "xls": ["application/vnd.ms-excel"],
        "xlsx": [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ],
    }
    signatures = {
        "json": [b"[", b" [", b"{", b" {"],
        "geojson": [b"[", b" [", b"{", b" {"],
        "shp": [b"PK\x03\x04"],
        "xls": [b"\xd0\xcf\x11\xe0"],
        "xlsx": [b"PK\x03\x04"],
    }

    def __init__(
        self, user_agent: str, url_ignore: Optional[str] = None
    ) -> None:
        self.user_agent = user_agent
        self.url_ignore: Optional[str] = url_ignore

    async def fetch(
        self,
        metadata: Tuple,
        session: Union[aiohttp.ClientSession, RateLimiter],
    ) -> Tuple:
        """Asynchronous code to download a resource and hash it. Returns a tuple with
        resource information including hashes.

        Args:
            metadata (Tuple): Resource to be checked
            session (Union[aiohttp.ClientSession, RateLimiter]): session to use for requests

        Returns:
            Tuple: Resource information including hash
        """
        url = metadata[0]
        resource_id = metadata[1]
        resource_format = metadata[2]

        async def fn(response):
            last_modified_str = response.headers.get("Last-Modified")
            http_last_modified = None
            if last_modified_str:
                try:
                    # we set http_last_modified but don't actually use it to calculate
                    # freshness any more
                    http_last_modified = parser.parse(
                        last_modified_str, ignoretz=True
                    )
                except (ValueError, OverflowError):
                    pass
            length = response.headers.get("Content-Length")
            if length and int(length) > 419430400:
                response.close()
                err = self.toolargeerror
                return (
                    resource_id,
                    url,
                    resource_format,
                    err,
                    http_last_modified,
                    None,
                    None,
                )
            logger.info(f"Hashing {url}")
            mimetype = response.headers.get("Content-Type")

            try:
                iterator = response.content.iter_any()
                first_chunk = await iterator.__anext__()
                signature = first_chunk[:4]
                if (
                    resource_format == "xlsx"
                    and mimetype == self.mimetypes["xlsx"][0]
                    and signature == self.signatures["xlsx"][0]
                    and (
                        self.url_ignore not in url if self.url_ignore else True
                    )
                ):
                    xlsxbuffer = bytearray(first_chunk)
                else:
                    xlsxbuffer = None
                md5hash = hashlib.md5(first_chunk)
                async for chunk in iterator:
                    if chunk:
                        md5hash.update(chunk)
                        if xlsxbuffer:
                            xlsxbuffer.extend(chunk)
                if xlsxbuffer:
                    workbook = load_workbook(
                        filename=BytesIO(xlsxbuffer), read_only=True
                    )
                    xlsx_md5hash = hashlib.md5()
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for cols in sheet.iter_rows(values_only=True):
                            xlsx_md5hash.update(bytes(str(cols), "utf-8"))
                    workbook.close()
                    xlsxbuffer = None
                else:
                    xlsx_md5hash = None
                err = None
                if mimetype not in self.ignore_mimetypes:
                    expected_mimetypes = self.mimetypes.get(resource_format)
                    if expected_mimetypes is not None:
                        if not any(x in mimetype for x in expected_mimetypes):
                            err = f"File mimetype {mimetype} {self.notmatcherror} {resource_format}!"
                expected_signatures = self.signatures.get(resource_format)
                if expected_signatures is not None:
                    found = False
                    for expected_signature in expected_signatures:
                        if (
                            signature[: len(expected_signature)]
                            == expected_signature
                        ):
                            found = True
                            break
                    if not found:
                        sigerr = f"File signature {signature} {self.notmatcherror} {resource_format}!"
                        if err is None:
                            err = sigerr
                        else:
                            err = f"{err} {sigerr}"
                return (
                    resource_id,
                    url,
                    resource_format,
                    err,
                    http_last_modified,
                    md5hash.hexdigest(),
                    xlsx_md5hash.hexdigest() if xlsx_md5hash else None,
                )
            except Exception as exc:
                try:
                    code = exc.code
                except AttributeError:
                    code = ""
                err = f"Exception during hashing: code={code} message={exc} raised={exc.__class__.__module__}.{exc.__class__.__qualname__} url={url}"
                raise aiohttp.ClientResponseError(
                    code=code,
                    message=err,
                    request_info=response.request_info,
                    history=response.history,
                ) from exc

        try:
            return await retry.send_http(
                session, "get", url, retries=2, interval=5, backoff=4, fn=fn
            )
        except Exception as e:
            return resource_id, url, resource_format, str(e), None, None, None

    async def check_urls(
        self, resources_to_check: List[Tuple], loop: uvloop.Loop
    ) -> Dict[str, Tuple]:
        """Asynchronous code to download resources and hash them. Return dictionary with
        resources information including hashes.

        Args:
            resources_to_check (List[Tuple]): List of resources to be checked
            loop (uvloop.Loop): Event loop to use

        Returns:
            Dict[str, Tuple]: Resources information including hashes
        """
        tasks = list()

        conn = aiohttp.TCPConnector(limit=100, limit_per_host=1, loop=loop)
        timeout = aiohttp.ClientTimeout(
            total=60 * 60, sock_connect=30, sock_read=30
        )
        async with aiohttp.ClientSession(
            connector=conn,
            timeout=timeout,
            loop=loop,
            headers={"User-Agent": self.user_agent},
        ) as session:
            session = RateLimiter(
                session
            )  # Limit connections per timeframe to host
            for metadata in resources_to_check:
                task = self.fetch(metadata, session)
                tasks.append(task)
            responses = dict()
            for f in tqdm.tqdm(asyncio.as_completed(tasks), total=len(tasks)):
                (
                    resource_id,
                    url,
                    resource_format,
                    err,
                    http_last_modified,
                    hash,
                    hash_xlsx,
                ) = await f
                responses[resource_id] = (
                    url,
                    resource_format,
                    err,
                    http_last_modified,
                    hash,
                    hash_xlsx,
                )
            return responses

    def retrieve(self, resources_to_check: List[Tuple]) -> Dict[str, Tuple]:
        """Download resources and hash them. Return dictionary with resources information
        including hashes.

        Args:
            resources_to_check (List[Tuple]): List of resources to be checked

        Returns:
            Dict[str, Tuple]: Resources information including hashes
        """

        start_time = timer()
        loop = uvloop.new_event_loop()
        asyncio.set_event_loop(loop)
        future = asyncio.ensure_future(
            self.check_urls(resources_to_check, loop)
        )
        results = loop.run_until_complete(future)
        logger.info(f"Execution time: {timer() - start_time} seconds")
        loop.run_until_complete(asyncio.sleep(0.250))
        loop.close()
        return results
